"""
分批生成 CentralizedBackup 测试用例 Excel v2
从 API 客户端模块源码自动提取接口信息
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import sys, os, importlib, inspect, re

sys.path.insert(0, os.path.dirname(__file__))
DST = "/Volume1/eee/cb-api-test/CentralizedBackup_API_TestCases.xlsx"

# ── 样式 ──
HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
CELL_FONT = Font(name="Arial", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

HEADERS = [
    "用例ID", "所属模块", "接口地址", "关联功能",
    "优先级", "接口请求类型", "用例标题",
    "前置条件", "操作步骤", "实际结果",
    "预期结果", "备注", "cURL",
]

MODULES = [
    ("PC/Server", "client.win_api", "WinAPI", "/v2/proxy/CentralizedBackup"),
    ("FileServer", "client.file_api", "FileAPI", "/v2/proxy2/CentralizedBackup"),
    ("VMware", "client.vmware_api", "VMwareAPI", "/v2/proxy/CentralizedBackup"),
    ("Hyper-V", "client.hyperv_api", "HyperVAPI", "/v2/proxy/CentralizedBackup"),
    ("TOS系统", "client.tos_api", "TOSAPI", ""),
]

# ── 工具函数 ──
TITLE_MAP = {
    "list": "查询列表", "get": "查询", "add": "添加", "create": "创建",
    "delete": "删除", "edit": "编辑", "start": "启动", "stop": "停止",
    "exec": "执行", "cancel": "取消", "lock": "锁定", "unlock": "解锁",
    "import": "导入", "upgrade": "升级", "browse": "浏览", "read": "读取",
    "set": "设置", "check": "检查", "download": "下载", "overview": "概览",
    "devices": "设备", "device": "设备", "task": "任务", "tasks": "任务",
    "version": "版本", "ver": "版本", "restore": "还原", "res": "还原",
    "login": "登录", "logins": "登录", "dir": "目录", "parts": "分区",
    "partitions": "分区", "disk": "磁盘", "disks": "磁盘",
    "state": "状态", "status": "状态", "detail": "详情",
    "verbose": "详细信息", "summary": "摘要",
    "vms": "虚拟机", "vm": "虚拟机", "linked": "关联",
    "config": "配置", "default": "默认", "letter": "盘符",
    "system": "系统", "image": "镜像", "boot": "启动镜像",
    "drv": "驱动器", "folder": "文件夹", "folders": "文件夹",
    "datacenter": "数据中心", "clusters": "集群", "cluster": "集群",
    "hosts": "主机", "host": "主机", "datastores": "数据存储",
    "networks": "网络", "restoretask": "还原任务",
    "sub": "子项", "batch": "批量", "relink": "重连",
    "guide": "引导", "locked": "锁定状态", "view": "视图",
    "smb": "SMB", "notification": "通知",
    "package": "安装包", "lang": "语言", "desktop": "桌面",
    "display": "显示", "info": "信息",
    "home": "首页", "volumes": "卷", "reinstall": "重装",
    "exist": "存在", "support": "支持", "enough": "充足",
    "resource": "资源",
}

def cn_title(name):
    return "".join(TITLE_MAP.get(p.lower(), p) for p in name.split("_") if p)

def guess_priority(method, path):
    if method == "GET" and "{id}" not in path and "{" not in path:
        return "P1"
    if method in ("POST",) and any(kw in path.lower() for kw in ["add", "create", "start"]):
        return "P1"
    if method == "PUT" and any(kw in path.lower() for kw in ["exec", "start"]):
        return "P1"
    if "check" in path.lower() or "browse" in path.lower() or "dir" in path.lower():
        return "P3"
    return "P2"

def guess_feature(name, path):
    n = name.lower()
    if any(kw in n for kw in ["device", "vm_list", "datacenter", "cluster", "host", "folder", "datastore", "network", "drv"]):
        return "设备管理"
    if any(kw in n for kw in ["restore", "res"]) and "task" not in n:
        return "还原管理"
    if any(kw in n for kw in ["restore_task", "restoretask"]):
        return "还原任务"
    if any(kw in n for kw in ["task", "exec", "cancel"]):
        return "任务管理"
    if any(kw in n for kw in ["version", "ver", "lock"]):
        return "版本管理"
    if any(kw in n for kw in ["login", "lang", "state"]):
        return "登录管理"
    if any(kw in n for kw in ["portal", "browse", "dir", "read", "list_dir", "disk_part", "drv"]):
        return "文件浏览"
    if "overview" in n:
        return "概览"
    if any(kw in n for kw in ["download", "client", "boot", "package"]):
        return "下载/安装"
    if any(kw in n for kw in ["folder_info", "home", "volume", "folder_list"]):
        return "文件/存储"
    return "其他"

def make_pre(path):
    if "{device_id}" in path or "{did}" in path:
        return "设备已存在，已登录TOS"
    if "{task_id}" in path or "{tid}" in path:
        return "任务已存在，已登录TOS"
    if "restore" in path.lower():
        return "已有完成的备份版本，已登录TOS"
    return "已登录TOS"

def make_expected(method, name):
    n = name.lower()
    if any(kw in n for kw in ["list", "get", "info", "state", "status", "detail", "verbose", "summary", "overview"]):
        return "返回200，data包含查询结果"
    if any(kw in n for kw in ["add", "create"]):
        return "返回200，创建成功"
    if "delete" in n:
        return "返回200，删除成功"
    if any(kw in n for kw in ["start", "exec"]):
        return "返回200，操作已开始执行"
    if any(kw in n for kw in ["stop", "cancel"]):
        return "返回200，操作已停止/取消"
    if any(kw in n for kw in ["edit", "set", "lock", "import"]):
        return "返回200，修改/操作成功"
    if "download" in n:
        return "返回200或文件流"
    if "check" in n:
        return "返回200，data包含检查结果"
    return "返回200，请求成功"

# ── 提取接口 ──
def extract_module(module_name, import_path, class_name, prefix):
    mod = importlib.import_module(import_path)
    cls = getattr(mod, class_name)
    try:
        src = inspect.getsource(cls)
    except (OSError, TypeError):
        src = ""

    # 从源码提取所有 self.client.get/post/put/delete 调用
    # 同时关联到方法名
    results = []
    for name, method in inspect.getmembers(cls, predicate=inspect.isfunction):
        if name.startswith("_"):
            continue
        try:
            method_src = inspect.getsource(method)
        except (OSError, TypeError):
            continue

        # 找所有HTTP调用
        http_calls = re.findall(r'self\.client\.(get|post|put|delete)\(\s*[fr]?["\x27]([^"\x27]+)["\x27]', method_src)
        if not http_calls:
            # 也可能是 client.get(f"...") 格式
            http_calls = re.findall(r'self\.client\.(get|post|put|delete)\(\s*f["\x27]([^"\x27]+)["\x27]', method_src)

        if not http_calls:
            continue

        for http_method, api_path in http_calls:
            # 对模板字符串做简化
            api_path = api_path.replace("{", "").replace("}", "")
            # 去掉 f-string 中的 {var} 但保留结构性路径
            api_path_clean = re.sub(r'\{[^}]+\}', lambda m: m.group(0), api_path)

            full_path = f"{prefix}{api_path}" if prefix else api_path
            pri = guess_priority(http_method.upper(), api_path)
            feat = guess_feature(name, api_path)
            title = cn_title(name)
            pre = make_pre(api_path)
            steps = f"1. {http_method.upper()} {api_path}"
            if http_method.upper() in ("POST", "PUT"):
                steps += "\n2. 传入必要请求体参数\n3. 检查响应"
            exp = make_expected(http_method, name)
            rmk = ""
            if api_path.count("/") >= 6:
                rmk = "深层嵌套路由"
            if "legacy" in name.lower() or "import" in name.lower():
                rmk = "兼容/特殊接口"

            results.append({
                "module": module_name, "path": full_path, "feature": feat,
                "priority": pri, "method": http_method.upper(), "title": title,
                "preconditions": pre, "steps": steps, "expected": exp, "remark": rmk,
            })

    # 去重（同名同路径）
    seen = set()
    unique = []
    for r in results:
        key = (r["module"], r["path"], r["method"])
        if key not in seen:
            seen.add(key)
            unique.append(r)
    return unique

def main():
    print("=== CentralizedBackup 测试用例 Excel 生成器 v2 ===")
    all_apis = []
    for module_name, import_path, class_name, prefix in MODULES:
        print(f"提取模块: {module_name}...")
        apis = extract_module(module_name, import_path, class_name, prefix)
        print(f"  → {len(apis)} 个接口")
        all_apis.extend(apis)
    print(f"\n总计: {len(all_apis)} 个接口\n")

    # 创建 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "API测试用例"

    # 写表头
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = THIN_BORDER

    # 列宽
    widths = [10, 12, 48, 12, 8, 12, 25, 22, 32, 12, 25, 16, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 分写入数据
    BATCH = 20
    row = 2
    for i, api in enumerate(all_apis):
        cid = i + 1
        data = [
            f"CB-{cid:03d}", api["module"], api["path"], api["feature"],
            api["priority"], api["method"], api["title"],
            api["preconditions"], api["steps"], "",
            api["expected"], api["remark"], "",
        ]
        for col, val in enumerate(data, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = CELL_FONT
            c.border = THIN_BORDER
            c.alignment = CENTER if col in (1,2,4,5,6) else LEFT_WRAP
        row += 1
        if cid % BATCH == 0:
            print(f"  已写入 {cid}/{len(all_apis)}...")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{row-1}"

    print(f"保存到 {DST}...")
    wb.save(DST)
    print(f"✅ 完成！共 {len(all_apis)} 条用例")

if __name__ == "__main__":
    main()
