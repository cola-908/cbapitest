#!/usr/bin/env python3
"""根据代码(client/*.py)重建 CentralizedBackup_API_TestCases.xlsx 接口文档"""
import ast, re, openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

PROXY_RULES = {
    "/win":    {"prefix": "/v2/proxy/CentralizedBackup",  "strip": "/win"},
    "/file":   {"prefix": "/v2/proxy2/CentralizedBackup", "strip": "/file"},
    "/VMware": {"prefix": "/v2/proxy2/CentralizedBackup", "strip": "/VMware"},
    "/hyperV": {"prefix": "/v2/proxy2/CentralizedBackup/58200", "strip": "/hyperV"},
}

def code_to_full(short):
    for pfx, rule in PROXY_RULES.items():
        if short.startswith(pfx):
            return rule["prefix"] + short.replace(rule["strip"], "", 1)
    return short

def parse_file(fp, mod):
    with open(fp) as f:
        src = f.read()
    tree = ast.parse(src)
    res = []
    for node in ast.walk(tree):
        if not isinstance(node, ast.FunctionDef) or node.name.startswith('_'):
            continue
        doc = ast.get_docstring(node) or ''
        found = False
        for m in ['GET','POST','PUT','DELETE','PATCH']:
            if doc.startswith(m+' '):
                path = doc[len(m):].strip().split()[0]
                res.append((mod, node.name, m, path))
                found = True
                break
        if not found:
            # Fallback: extract from self.client.{get,post,put,delete}(path, ...)
            for child in ast.walk(node):
                if isinstance(child, ast.Call):
                    if (isinstance(child.func, ast.Attribute) and
                        isinstance(child.func.value, ast.Attribute) and
                        child.func.value.attr == 'client' and
                        child.func.attr in ('get','post','put','delete','patch')):
                        method = child.func.attr.upper()
                        if child.args:
                            arg = child.args[0]
                            if isinstance(arg, ast.Constant):
                                path = arg.value
                                res.append((mod, node.name, method, path))
                                break
                            elif isinstance(arg, ast.JoinedStr):
                                # Reconstruct f-string path
                                parts = []
                                for v in arg.values:
                                    if isinstance(v, ast.Constant):
                                        parts.append(v.value)
                                    elif isinstance(v, ast.FormattedValue):
                                        parts.append('{%s}' % (
                                            v.value.id if isinstance(v.value, ast.Name) else 'param'))
                                path = ''.join(parts)
                                res.append((mod, node.name, method, path))
                                break
    return [r for r in res if r[1] != '__init__']

apis = []
apis += parse_file('client/win_api.py', 'PC/Server')
apis += parse_file('client/file_api.py', 'FileServer')
apis += parse_file('client/vmware_api.py', 'VMware')
apis += parse_file('client/hyperv_api.py', 'Hyper-V')
apis += parse_file('client/tos_api.py', 'TOS系统')

TITLE = {
    'list_devices':'查询列表设备','get_device':'查询设备','list_all_devices':'查询列表all设备',
    'delete_device':'删除设备','batch_delete_devices':'批量删除设备','get_device_volumes':'查询设备卷',
    'upgrade_device':'升级设备','get_device_tasks_summary':'查询设备任务摘要',
    'list_tasks':'查询列表任务','get_task':'查询任务','list_all_tasks':'查询列表all任务',
    'exec_task':'执行任务','cancel_task':'取消任务','get_task_verbose':'查询任务详细信息',
    'edit_task':'编辑任务','relink_task':'重连任务',
    'batch_add_tasks':'批量添加任务','batch_exec_tasks':'批量执行任务',
    'batch_cancel_tasks':'批量取消任务','batch_delete_tasks':'批量删除任务','batch_edit_tasks':'批量编辑任务',
    'list_versions':'查询列表versions','get_version':'查询版本','get_version_guide':'查询版本引导',
    'get_version_detail':'查询版本详情','lock_version':'锁定版本','get_disk_partitions':'查询磁盘分区',
    'list_restore_tasks':'查询列表还原任务','create_restore_task':'创建还原任务',
    'cancel_restore':'取消还原','delete_restore':'删除还原','restore':'还原',
    'browse_image':'浏览镜像','get_dev_dir_sub_dirs':'查询dev目录子项dirs',
    'portal_dir':'portal目录','download_pc_client':'下载pcclient','download_server_client':'下载serverclient',
    'overview':'概览','get_task_detail':'查询任务详情','get_device_task_versions':'查询设备任务versions',
    'add_device':'添加设备','read_dir':'读取目录',
    'list_logins':'查询列表登录','add_login':'添加登录','edit_login':'编辑登录','get_login_state':'查询登录状态',
    'list_backup_tasks':'查询列表backup任务','get_backup_task':'查询backup任务',
    'add_backup_task':'添加backup任务','edit_backup_task':'编辑backup任务',
    'start_backup_task':'启动backup任务','stop_backup_task':'停止backup任务',
    'start_backup':'启动backup','stop_backup':'停止backup','import_backup':'导入backup',
    'delete_version':'删除版本','set_version_locked':'设置版本锁定状态',
    'delete_backup_version':'删除backup版本','lock_backup_version':'锁定backup版本',
    'get_restore_task':'查询还原任务','list_restores':'查询列表restores',
    'get_restore':'查询还原','start_restore':'启动还原','stop_restore':'停止还原',
    'start_restore_task':'启动还原任务','stop_restore_task':'停止还原任务',
    'list_tasks_legacy':'查询列表任务legacy','get_smb_info':'查询SMB信息',
    'check_resource_enough':'检查资源充足','check_restore_support':'检查还原支持',
    'check_vm_exist':'检查虚拟机存在','get_device_status':'查询设备状态',
    'get_linked_task':'查询关联任务','get_task_devices':'查询任务设备',
    'list_datacenters':'查询列表datacenters','list_datacenter_clusters':'查询列表数据中心集群',
    'list_datacenter_cluster_hosts':'查询列表数据中心集群主机',
    'list_datacenter_cluster_host_vms':'查询列表数据中心集群主机虚拟机',
    'list_datacenter_folders':'查询列表数据中心文件夹',
    'list_datacenter_folder_vms':'查询列表数据中心文件夹虚拟机',
    'list_datacenter_datastores':'查询列表数据中心数据存储',
    'list_datacenter_networks':'查询列表数据中心网络',
    'list_datacenter_vms':'查询列表数据中心虚拟机','get_datacenter_vm':'查询数据中心虚拟机',
    'add_task':'添加任务','import_task':'导入任务','start_task':'启动任务','stop_task':'停止任务',
    'check_dest_folder':'检查dest文件夹','check_vm_service':'检查虚拟机service',
    'get_partition_system':'查询partition系统','get_task_ver_vms':'查询任务版本虚拟机',
    'set_ver_locked':'设置版本锁定状态',
    'start_restore_task':'启动还原任务','stop_restore_task':'停止还原任务',
    'delete_restore_task':'删除还原任务','vm_list':'虚拟机查询列表',
    'edit_device':'编辑设备','get_device_state':'查询设备状态',
    'get_default_cfg':'查询默认cfg','get_vm_list':'查询虚拟机列表',
    'list_device_vms':'查询列表设备虚拟机','get_vm_default_dirs':'查询虚拟机默认dirs',
    'list_device_drv_dir':'查询列表设备驱动器目录','list_device_drv_letter':'查询列表设备驱动器盘符',
    'list_dir':'查询列表目录','list_device_tasks':'查询列表设备任务',
    'delete_task_version':'删除任务版本','get_task_version_detail':'查询任务版本详情',
    'delete_task_sub':'删除任务子项',
    'add_restore':'添加还原','exec_restore':'执行还原',
    'stop_restore':'停止还原','delete_restore':'删除还原',
    'browse_part':'浏览part','disk_parts':'磁盘分区',
    'disk_view_browse':'磁盘视图浏览','get_disk_parts':'查询磁盘分区',
    'download_boot_image':'下载启动镜像',
    'get_lang':'查询语言','get_app_version':'查询app版本',
    'get_version_generic':'查询版本generic','get_desktop_display':'查询桌面显示',
    'create_folder':'创建文件夹','get_folder_info':'查询文件夹信息',
    'get_home_list':'查询首页查询列表','list_folders':'查询列表文件夹',
    'list_volumes':'查询列表卷','reinstall_system':'重装系统',
    'download_client_package':'下载client安装包',
}

FUNC_GROUPS = {
    'PC/Server': {
        'list_devices':'设备管理','get_device':'设备管理','list_all_devices':'设备管理',
        'delete_device':'设备管理','batch_delete_devices':'设备管理',
        'get_device_volumes':'设备管理','upgrade_device':'设备管理','get_device_tasks_summary':'设备管理',
        'list_tasks':'任务管理','get_task':'任务管理','list_all_tasks':'任务管理',
        'exec_task':'任务管理','cancel_task':'任务管理','get_task_verbose':'任务管理',
        'edit_task':'任务管理','relink_task':'任务管理',
        'batch_add_tasks':'任务管理','batch_exec_tasks':'任务管理',
        'batch_cancel_tasks':'任务管理','batch_delete_tasks':'任务管理','batch_edit_tasks':'任务管理',
        'list_versions':'版本管理','get_version':'版本管理','get_version_guide':'版本管理',
        'get_version_detail':'版本管理','lock_version':'版本管理','get_disk_partitions':'版本管理',
        'list_restore_tasks':'还原','create_restore_task':'还原',
        'cancel_restore':'还原','delete_restore':'还原','restore':'还原',
        'browse_image':'Portal浏览','get_dev_dir_sub_dirs':'Portal浏览',
        'portal_dir':'Portal浏览','download_pc_client':'Portal浏览','download_server_client':'Portal浏览',
        'overview':'概览','get_task_detail':'任务管理','get_device_task_versions':'版本管理',
    },
    'FileServer': {
        'list_devices':'设备管理','get_device':'设备管理','add_device':'设备管理',
        'delete_device':'设备管理','read_dir':'设备管理',
        'list_logins':'登录管理','add_login':'登录管理','edit_login':'登录管理','get_login_state':'登录管理',
        'list_backup_tasks':'备份任务','get_backup_task':'备份任务','add_backup_task':'备份任务',
        'edit_backup_task':'备份任务','start_backup_task':'备份任务','stop_backup_task':'备份任务',
        'start_backup':'备份任务','stop_backup':'备份任务','import_backup':'备份任务',
        'list_versions':'版本管理','get_version':'版本管理','delete_version':'版本管理',
        'set_version_locked':'版本管理','delete_backup_version':'版本管理','lock_backup_version':'版本管理',
        'list_restore_tasks':'还原任务','get_restore_task':'还原任务',
        'list_restores':'还原任务','get_restore':'还原任务',
        'start_restore':'还原任务','stop_restore':'还原任务',
        'start_restore_task':'还原任务','stop_restore_task':'还原任务',
        'list_tasks_legacy':'其他','get_smb_info':'其他','overview':'概览',
    },
    'VMware': {
        'list_devices':'设备管理','get_device':'设备管理','add_device':'设备管理',
        'delete_device':'设备管理','check_resource_enough':'设备管理',
        'check_restore_support':'设备管理','check_vm_exist':'设备管理',
        'get_device_status':'设备管理','get_linked_task':'设备管理','get_task_devices':'设备管理',
        'list_datacenters':'数据中心浏览','list_datacenter_clusters':'数据中心浏览',
        'list_datacenter_cluster_hosts':'数据中心浏览','list_datacenter_cluster_host_vms':'数据中心浏览',
        'list_datacenter_folders':'数据中心浏览','list_datacenter_folder_vms':'数据中心浏览',
        'list_datacenter_datastores':'数据中心浏览','list_datacenter_networks':'数据中心浏览',
        'list_datacenter_vms':'数据中心浏览','get_datacenter_vm':'数据中心浏览',
        'list_tasks':'任务管理','get_task':'任务管理','add_task':'任务管理',
        'import_task':'任务管理','start_task':'任务管理','stop_task':'任务管理',
        'check_dest_folder':'任务管理','check_vm_service':'任务管理',
        'get_disk_partitions':'任务管理','get_partition_system':'任务管理','get_task_ver_vms':'任务管理',
        'list_versions':'版本管理','get_version':'版本管理','set_ver_locked':'版本管理',
        'list_restore_tasks':'还原任务','get_restore_task':'还原任务',
        'start_restore_task':'还原任务','stop_restore_task':'还原任务','delete_restore_task':'还原任务',
        'vm_list':'虚拟机查询','overview':'概览',
    },
    'Hyper-V': {
        'list_devices':'设备管理','get_device':'设备管理','add_device':'设备管理',
        'edit_device':'设备管理','delete_device':'设备管理',
        'get_device_state':'设备管理','get_default_cfg':'设备管理',
        'get_vm_list':'设备管理','list_device_vms':'设备管理',
        'get_vm_default_dirs':'设备管理','list_device_drv_dir':'设备管理',
        'list_device_drv_letter':'设备管理','list_dir':'设备管理','list_device_tasks':'设备管理',
        'list_tasks':'任务管理','get_task':'任务管理','get_task_detail':'任务管理',
        'add_task':'任务管理','edit_task':'任务管理','exec_task':'任务管理',
        'stop_task':'任务管理','delete_task':'任务管理',
        'delete_task_version':'任务管理','get_task_version_detail':'任务管理','delete_task_sub':'任务管理',
        'list_restore_tasks':'还原任务','get_restore_task':'还原任务',
        'list_restores':'还原任务','get_restore':'还原任务',
        'add_restore':'还原任务','exec_restore':'还原任务',
        'stop_restore':'还原任务','delete_restore':'还原任务',
        'browse_part':'Portal浏览','disk_parts':'Portal浏览',
        'disk_view_browse':'Portal浏览','get_disk_parts':'Portal浏览',
        'overview':'概览','download_boot_image':'下载',
    },
    'TOS系统': {
        'get_lang':'语言','get_app_version':'应用版本','get_version_generic':'应用版本',
        'get_desktop_display':'桌面显示','create_folder':'文件管理',
        'get_folder_info':'文件管理','get_home_list':'文件管理',
        'list_folders':'文件管理','list_volumes':'存储卷',
        'reinstall_system':'系统恢复','download_client_package':'客户端下载',
    },
}

PRIORITY = {
    '设备管理':'P1','任务管理':'P1','概览':'P1','还原任务':'P1','还原':'P1','备份任务':'P1','系统恢复':'P1',
    '版本管理':'P2','登录管理':'P2','Portal浏览':'P2','数据中心浏览':'P2','虚拟机查询':'P2','文件管理':'P2','存储卷':'P2','客户端下载':'P2','下载':'P2',
    '语言':'P3','应用版本':'P3','桌面显示':'P3','其他':'P3',
}

# Read old xlsx for preserved step/expected data
try:
    old_wb = openpyxl.load_workbook('CentralizedBackup_API_TestCases.xlsx')
    old_ws = old_wb['API测试用例']
except:
    old_wb = None
    old_ws = None

old_info = {}
if old_ws:
    for row in old_ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None: continue
        api_url = row[2]; method = row[5]
        norm = api_url
        for pfx in ['/v2/proxy2/CentralizedBackup/58200','/v2/proxy2/CentralizedBackup','/v2/proxy/CentralizedBackup']:
            if norm.startswith(pfx):
                norm = norm[len(pfx):]; break
        norm = re.sub(r'\{[^}]+\}','{}',norm)
        norm = re.sub(r'(?<=/)(device_id|task_id|ver_id|sub_id|detail_id|app_id|login_id|restore_id|vm_name|disk_id|part_id|file_path)(?=/|$)','{}',norm)
        steps = row[8] if len(row) > 8 else ''
        expected = row[10] if len(row) > 10 else ''
        title = row[6] if row[6] else ''
        old_info[(method, norm)] = {'steps': steps, 'expected': expected, 'title': title}

# Build entries
entries = []
idx = 0
for mod, func, method, short in apis:
    idx += 1
    full = code_to_full(short)
    fg = FUNC_GROUPS.get(mod, {}).get(func, '其他')
    pri = PRIORITY.get(fg, 'P3')
    title = TITLE.get(func, func)
    
    # Match old data by biz path
    norm_full = full
    for pfx in ['/v2/proxy2/CentralizedBackup/58200','/v2/proxy2/CentralizedBackup','/v2/proxy/CentralizedBackup']:
        if norm_full.startswith(pfx):
            norm_full = norm_full[len(pfx):]; break
    norm_full = re.sub(r'\{[^}]+\}','{}',norm_full)
    
    old = old_info.get((method, norm_full), {})
    # Prefer old title if it exists and is Chinese
    old_title = old.get('title','')
    if old_title and re.search(r'[\u4e00-\u9fff]', old_title):
        title = old_title
    
    steps = old.get('steps','') or ("1. %s %s\n2. 传入必要参数\n3. 检查响应" % (method, full))
    expected = old.get('expected','') or ('返回200，操作成功' if method != 'GET' else '返回200，查询成功')
    
    entries.append([("CB-%03d" % idx), mod, full, fg, pri, method, title, '已登录TOS', steps, None, expected, None, None])

# Write
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'API测试用例'

hdr = ['用例ID','所属模块','接口地址','关联功能','优先级','接口请求类型','用例标题','前置条件','操作步骤','实际结果','预期结果','备注','cURL']
hfont = Font(bold=True, size=11, color='FFFFFF')
hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
bdr = Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))

for c, h in enumerate(hdr, 1):
    cell = ws.cell(1, c, h)
    cell.font = hfont; cell.fill = hfill
    cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
    cell.border = bdr

for r, row in enumerate(entries, 2):
    for c, val in enumerate(row, 1):
        cell = ws.cell(r, c, val)
        cell.alignment = Alignment(vertical='top',wrap_text=True)
        cell.border = bdr

widths = [10,12,65,14,8,10,22,12,40,12,20,10,15]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

ws.freeze_panes = 'A2'
wb.save('CentralizedBackup_API_TestCases.xlsx')
print("Done! Wrote %d entries" % len(entries))
